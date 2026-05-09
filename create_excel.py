"""
NVIDIA NPN 파트너 관리 기획서 v2
시트 1: 1.3 FY별 Competency 매출 추이 (CAGR + FY Mix% 자동 계산)
시트 2: 파트너 × Competency 히트맵 (Revenue 입력 → 조건부 서식 자동)
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
import os

OUT_FILE = "/Users/sangokh/Desktop/기획서/NVIDIA_NPN_파트너_관리_기획서_v2.xlsx"

# ── 색상 팔레트 ────────────────────────────────────────────────────────────────
NVIDIA_GREEN  = "76B900"
NAVY_DARK     = "1A1A2E"
HEADER_FG     = "FFFFFF"
GRAY_LIGHT    = "F5F5F5"
GRAY_MID      = "E0E0E0"
ACCENT_BLUE   = "D5E8F0"
ACCENT_GREEN  = "D5E8D4"
YELLOW_WARN   = "FFF2CC"
TEXT_DARK     = "333333"

# Competency 목록
COMPETENCIES = ["AI-COMP", "AI-ENT", "DC-COMP", "NETWORK", "VIZ", "VIRT", "EDGE", "AUTO"]

# 파트너 예시 (Type 포함)
PARTNERS = [
    ("Partner A", "SP"),
    ("Partner B", "SP"),
    ("Partner C", "SP"),
    ("Partner D", "SP"),
    ("Partner E", "SA"),
    ("Partner F", "SA"),
    ("Partner G", "Distributor"),
    ("Partner H", "Distributor"),
]

FY_YEARS = ["FY24", "FY25", "FY26"]


# ── 공통 스타일 헬퍼 ──────────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border(sides=("top","bottom","left","right")):
    s = Side(style="thin", color="CCCCCC")
    kw = {k: s for k in sides}
    return Border(**kw)

def all_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(top=s, bottom=s, left=s, right=s)

def bold_font(size=10, color=TEXT_DARK, name="맑은 고딕"):
    return Font(bold=True, size=size, color=color, name=name)

def normal_font(size=10, color=TEXT_DARK, name="맑은 고딕"):
    return Font(size=size, color=color, name=name)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")


# ══════════════════════════════════════════════════════════════════════════════
# 시트 1: FY별 Competency 매출 추이
# ══════════════════════════════════════════════════════════════════════════════
def build_revenue_sheet(wb):
    ws = wb.create_sheet("1.3_매출추이분석")
    ws.sheet_view.showGridLines = False

    # ── 컬럼 너비 설정 ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16   # Competency
    ws.column_dimensions["C"].width = 14   # FY24
    ws.column_dimensions["D"].width = 14   # FY25
    ws.column_dimensions["E"].width = 14   # FY26
    ws.column_dimensions["F"].width = 12   # CAGR
    ws.column_dimensions["G"].width = 13   # FY24 Mix%
    ws.column_dimensions["H"].width = 13   # FY25 Mix%
    ws.column_dimensions["I"].width = 13   # FY26 Mix%
    ws.column_dimensions["J"].width = 5

    # ── 타이틀 ────────────────────────────────────────────────────────────────
    ws.merge_cells("B1:I1")
    t = ws["B1"]
    t.value = "1.3  FY별 Competency 매출 추이 분석"
    t.font      = Font(bold=True, size=14, color=NVIDIA_GREEN, name="맑은 고딕")
    t.alignment = left_align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("B2:I2")
    sub = ws["B2"]
    sub.value = "※ Revenue(USD) 입력 → CAGR 및 FY Mix%가 자동 계산됩니다.  |  기준: Net Revenue (USD)"
    sub.font      = Font(size=9, color="888888", italic=True, name="맑은 고딕")
    sub.alignment = left_align()
    ws.row_dimensions[2].height = 16

    # ── 헤더 행 (행 4) ────────────────────────────────────────────────────────
    headers = ["Competency", "FY24\nRevenue (USD)", "FY25\nRevenue (USD)",
               "FY26\nRevenue (USD)", "CAGR\n(FY24→26)", "FY24\nMix%", "FY25\nMix%", "FY26\nMix%"]
    cols    = ["B","C","D","E","F","G","H","I"]

    ws.row_dimensions[4].height = 36
    for col, hdr in zip(cols, headers):
        cell = ws[f"{col}4"]
        cell.value     = hdr
        cell.font      = bold_font(10, HEADER_FG)
        cell.fill      = hdr_fill(NAVY_DARK)
        cell.alignment = center_align(wrap=True)
        cell.border    = all_border()

    # ── 데이터 행 (행 5 ~ 12) ─────────────────────────────────────────────────
    total_fy24_cell = f"C{5 + len(COMPETENCIES)}"  # 합계행 FY24
    total_fy25_cell = f"D{5 + len(COMPETENCIES)}"
    total_fy26_cell = f"E{5 + len(COMPETENCIES)}"

    for idx, comp in enumerate(COMPETENCIES):
        row = 5 + idx
        bg  = GRAY_LIGHT if idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[row].height = 22

        # Competency 이름
        c = ws[f"B{row}"]
        c.value     = comp
        c.font      = bold_font(10, TEXT_DARK)
        c.fill      = hdr_fill(bg)
        c.alignment = left_align()
        c.border    = all_border()

        # FY24 / FY25 / FY26 입력 셀
        for col in ["C","D","E"]:
            cell = ws[f"{col}{row}"]
            cell.value        = None       # 사용자 입력
            cell.number_format = '#,##0'
            cell.fill         = hdr_fill(ACCENT_BLUE if col == "E" else bg)
            cell.alignment    = right_align()
            cell.border       = all_border()
            cell.font         = normal_font(10, TEXT_DARK)
            if col in ("C","D"):
                # 연한 배경으로 입력 유도
                cell.fill = PatternFill("solid", fgColor="EAF3FB" if col == "D" else "F0F8FF")

        # CAGR = (FY26/FY24)^(1/2) - 1  [FY24→FY26 = 2년]
        # 분모 0 방지: IF(AND(C{row}>0, E{row}>0), ...)
        cagr_col = f"F{row}"
        ws[cagr_col].value = (
            f'=IF(AND(C{row}>0,E{row}>0),'
            f'POWER(E{row}/C{row},1/2)-1,"")'
        )
        ws[cagr_col].number_format = "0.0%"
        ws[cagr_col].fill          = hdr_fill(ACCENT_GREEN)
        ws[cagr_col].alignment     = center_align()
        ws[cagr_col].border        = all_border()
        ws[cagr_col].font          = normal_font(10, TEXT_DARK)

        # FY Mix% — 해당 FY 전체 합계 대비 비율
        for mix_col, fy_col, tot_col in [
            ("G", "C", "C"), ("H", "D", "D"), ("I", "E", "E")
        ]:
            tot_ref = f"{tot_col}{5 + len(COMPETENCIES)}"
            cell = ws[f"{mix_col}{row}"]
            cell.value = (
                f'=IF(AND({fy_col}{row}>0,{tot_ref}>0),'
                f'{fy_col}{row}/{tot_ref},"")'
            )
            cell.number_format = "0.0%"
            cell.fill          = hdr_fill(YELLOW_WARN if mix_col == "I" else bg)
            cell.alignment     = center_align()
            cell.border        = all_border()
            cell.font          = normal_font(10, TEXT_DARK)

    # ── 합계 행 ───────────────────────────────────────────────────────────────
    tot_row = 5 + len(COMPETENCIES)
    ws.row_dimensions[tot_row].height = 24
    data_range = f"5:{4 + len(COMPETENCIES)}"

    tot = ws[f"B{tot_row}"]
    tot.value     = "합계"
    tot.font      = bold_font(11, HEADER_FG)
    tot.fill      = hdr_fill(NAVY_DARK)
    tot.alignment = center_align()
    tot.border    = all_border()

    for col in ["C","D","E"]:
        cell = ws[f"{col}{tot_row}"]
        cell.value = (
            f"=SUM({col}5:{col}{4 + len(COMPETENCIES)})"
        )
        cell.number_format = '#,##0'
        cell.font          = bold_font(11, HEADER_FG)
        cell.fill          = hdr_fill(NAVY_DARK)
        cell.alignment     = right_align()
        cell.border        = all_border()

    # CAGR 합계 (전체 포트폴리오)
    ws[f"F{tot_row}"].value = (
        f'=IF(AND(C{tot_row}>0,E{tot_row}>0),'
        f'POWER(E{tot_row}/C{tot_row},1/2)-1,"")'
    )
    ws[f"F{tot_row}"].number_format = "0.0%"
    ws[f"F{tot_row}"].font          = bold_font(11, HEADER_FG)
    ws[f"F{tot_row}"].fill          = hdr_fill(NVIDIA_GREEN)
    ws[f"F{tot_row}"].alignment     = center_align()
    ws[f"F{tot_row}"].border        = all_border()

    for col in ["G","H","I"]:
        cell = ws[f"{col}{tot_row}"]
        cell.value     = "100%"
        cell.font      = bold_font(11, HEADER_FG)
        cell.fill      = hdr_fill(NAVY_DARK)
        cell.alignment = center_align()
        cell.border    = all_border()

    # ── 조건부 서식: CAGR 색상 ────────────────────────────────────────────────
    # CAGR > 20% → 진초록, 0~20% → 연초록, <0% → 연빨강
    ws.conditional_formatting.add(
        f"F5:F{4 + len(COMPETENCIES)}",
        ColorScaleRule(
            start_type="num", start_value=-0.3,
            start_color="FF9999",
            mid_type="num", mid_value=0,
            mid_color="FFFF99",
            end_type="num", end_value=0.5,
            end_color="63BE7B"
        )
    )

    # ── 조건부 서식: FY26 Mix% 크기 시각화 ───────────────────────────────────
    ws.conditional_formatting.add(
        f"I5:I{4 + len(COMPETENCIES)}",
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max", end_color=NVIDIA_GREEN
        )
    )

    # ── 안내 메모 ─────────────────────────────────────────────────────────────
    note_row = tot_row + 2
    ws.merge_cells(f"B{note_row}:I{note_row}")
    note = ws[f"B{note_row}"]
    note.value = (
        "💡 입력 방법: C~E열(파란 배경)에 각 Competency의 FY별 Net Revenue(USD)를 입력하세요. "
        "CAGR과 Mix%가 자동으로 계산됩니다."
    )
    note.font      = Font(size=9, color="555555", italic=True, name="맑은 고딕")
    note.alignment = left_align(wrap=True)
    ws.row_dimensions[note_row].height = 20


# ══════════════════════════════════════════════════════════════════════════════
# 시트 2: 파트너 × Competency 히트맵
# ══════════════════════════════════════════════════════════════════════════════
def build_heatmap_sheet(wb):
    ws = wb.create_sheet("히트맵_Revenue")
    ws.sheet_view.showGridLines = False

    N_PARTNERS = len(PARTNERS)
    N_COMP     = len(COMPETENCIES)

    # ── 컬럼 너비 ─────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16   # 파트너명
    ws.column_dimensions["C"].width = 8    # Type
    for i in range(N_COMP):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12
    ws.column_dimensions[get_column_letter(4 + N_COMP)].width = 13  # 합계
    ws.column_dimensions[get_column_letter(5 + N_COMP)].width = 3

    # ── 타이틀 ────────────────────────────────────────────────────────────────
    last_data_col = get_column_letter(4 + N_COMP)
    ws.merge_cells(f"B1:{last_data_col}1")
    t = ws["B1"]
    t.value = "파트너 × Competency Revenue 히트맵"
    t.font      = Font(bold=True, size=14, color=NVIDIA_GREEN, name="맑은 고딕")
    t.alignment = left_align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"B2:{last_data_col}2")
    sub = ws["B2"]
    sub.value = (
        "※ Revenue(USD) 입력 → 셀 색상이 자동으로 업데이트됩니다.  "
        "색이 진할수록 매출이 높습니다. (흰색=미입력, 연초록→진초록=저→고)"
    )
    sub.font      = Font(size=9, color="888888", italic=True, name="맑은 고딕")
    sub.alignment = left_align()
    ws.row_dimensions[2].height = 16

    # ── 헤더 행 (행 4) ────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 36

    # 파트너명
    for col_idx, (label, width_col) in enumerate([
        ("파트너명", "B"), ("Type", "C")
    ]):
        cell = ws[f"{width_col}4"]
        cell.value     = label
        cell.font      = bold_font(10, HEADER_FG)
        cell.fill      = hdr_fill(NAVY_DARK)
        cell.alignment = center_align()
        cell.border    = all_border()

    # Competency 헤더
    for j, comp in enumerate(COMPETENCIES):
        col = get_column_letter(4 + j)
        cell = ws[f"{col}4"]
        cell.value     = comp
        cell.font      = bold_font(10, HEADER_FG)
        cell.fill      = hdr_fill(NVIDIA_GREEN)
        cell.alignment = center_align(wrap=True)
        cell.border    = all_border()

    # 합계 헤더
    sum_col = get_column_letter(4 + N_COMP)
    cell = ws[f"{sum_col}4"]
    cell.value     = "파트너\n합계"
    cell.font      = bold_font(10, HEADER_FG)
    cell.fill      = hdr_fill(NAVY_DARK)
    cell.alignment = center_align(wrap=True)
    cell.border    = all_border()

    # ── 데이터 행 ─────────────────────────────────────────────────────────────
    data_start_row = 5
    data_end_row   = data_start_row + N_PARTNERS - 1

    for idx, (partner, ptype) in enumerate(PARTNERS):
        row = data_start_row + idx
        bg  = GRAY_LIGHT if idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[row].height = 22

        # 파트너명
        cell = ws[f"B{row}"]
        cell.value     = partner
        cell.font      = bold_font(10, TEXT_DARK)
        cell.fill      = hdr_fill(bg)
        cell.alignment = left_align()
        cell.border    = all_border()

        # Type
        cell = ws[f"C{row}"]
        type_color = {"SP": "D5E8D4", "SA": "DAE8FC", "Distributor": "FFE6CC"}.get(ptype, bg)
        cell.value     = ptype
        cell.font      = normal_font(9, TEXT_DARK)
        cell.fill      = hdr_fill(type_color)
        cell.alignment = center_align()
        cell.border    = all_border()

        # Revenue 입력 셀 (Competency별)
        rev_cells = []
        for j in range(N_COMP):
            col  = get_column_letter(4 + j)
            addr = f"{col}{row}"
            rev_cells.append(addr)
            cell = ws[addr]
            cell.value         = None
            cell.number_format = '#,##0'
            cell.fill          = hdr_fill("FAFFFE")
            cell.alignment     = right_align()
            cell.border        = all_border()
            cell.font          = normal_font(10, TEXT_DARK)

        # 파트너 합계
        sum_col_letter = get_column_letter(4 + N_COMP)
        first_rev_col  = get_column_letter(4)
        last_rev_col   = get_column_letter(3 + N_COMP)
        cell = ws[f"{sum_col_letter}{row}"]
        cell.value         = f"=SUM({first_rev_col}{row}:{last_rev_col}{row})"
        cell.number_format = '#,##0'
        cell.fill          = hdr_fill(ACCENT_BLUE)
        cell.alignment     = right_align()
        cell.border        = all_border()
        cell.font          = bold_font(10, TEXT_DARK)

    # ── Competency 합계 행 ────────────────────────────────────────────────────
    tot_row = data_end_row + 1
    ws.row_dimensions[tot_row].height = 24

    cell = ws[f"B{tot_row}"]
    cell.value     = "Competency 합계"
    cell.font      = bold_font(11, HEADER_FG)
    cell.fill      = hdr_fill(NAVY_DARK)
    cell.alignment = center_align()
    cell.border    = all_border()

    cell = ws[f"C{tot_row}"]
    cell.font  = bold_font(11, HEADER_FG)
    cell.fill  = hdr_fill(NAVY_DARK)
    cell.border = all_border()

    for j in range(N_COMP):
        col  = get_column_letter(4 + j)
        cell = ws[f"{col}{tot_row}"]
        cell.value         = f"=SUM({col}{data_start_row}:{col}{data_end_row})"
        cell.number_format = '#,##0'
        cell.font          = bold_font(11, HEADER_FG)
        cell.fill          = hdr_fill(NVIDIA_GREEN)
        cell.alignment     = right_align()
        cell.border        = all_border()

    # 전체 합계
    sum_col_letter = get_column_letter(4 + N_COMP)
    cell = ws[f"{sum_col_letter}{tot_row}"]
    cell.value         = (
        f"=SUM({get_column_letter(4)}{tot_row}:{get_column_letter(3+N_COMP)}{tot_row})"
    )
    cell.number_format = '#,##0'
    cell.font          = bold_font(11, HEADER_FG)
    cell.fill          = hdr_fill(NAVY_DARK)
    cell.alignment     = right_align()
    cell.border        = all_border()

    # ── 조건부 서식: 히트맵 (Revenue 입력 셀 전체) ────────────────────────────
    # 데이터 영역: D5 ~ (last_comp_col)(data_end_row)
    data_area = (
        f"{get_column_letter(4)}{data_start_row}:"
        f"{get_column_letter(3 + N_COMP)}{data_end_row}"
    )
    ws.conditional_formatting.add(
        data_area,
        ColorScaleRule(
            start_type="num",  start_value=0,       start_color="FFFFFF",
            mid_type="num",    mid_value=1000000,   mid_color="B8E4B8",
            end_type="max",                          end_color=NVIDIA_GREEN
        )
    )

    # 파트너 합계 열 조건부 서식
    sum_col_area = (
        f"{get_column_letter(4 + N_COMP)}{data_start_row}:"
        f"{get_column_letter(4 + N_COMP)}{data_end_row}"
    )
    ws.conditional_formatting.add(
        sum_col_area,
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="1F6B8E"
        )
    )

    # ── Type 범례 ─────────────────────────────────────────────────────────────
    legend_row = tot_row + 2
    ws.merge_cells(f"B{legend_row}:{get_column_letter(4 + N_COMP)}{legend_row}")
    ws[f"B{legend_row}"].value = (
        "Type 범례:   SP (연초록)     SA (연파랑)     Distributor (연주황)   "
        "│  히트맵: 흰색(0) → 연초록 → 진초록(최대값)  │  파트너합계: 흰색 → 파랑"
    )
    ws[f"B{legend_row}"].font      = Font(size=9, color="555555", italic=True, name="맑은 고딕")
    ws[f"B{legend_row}"].alignment = left_align()
    ws.row_dimensions[legend_row].height = 18

    note_row = legend_row + 1
    ws.merge_cells(f"B{note_row}:{get_column_letter(4 + N_COMP)}{note_row}")
    ws[f"B{note_row}"].value = (
        "💡 입력 방법: 각 파트너의 Competency별 FY26 Net Revenue(USD)를 흰색 셀에 입력하세요. "
        "파트너/Competency 합계와 히트맵 색상이 자동으로 업데이트됩니다."
    )
    ws[f"B{note_row}"].font      = Font(size=9, color="555555", italic=True, name="맑은 고딕")
    ws[f"B{note_row}"].alignment = left_align(wrap=True)
    ws.row_dimensions[note_row].height = 20


# ══════════════════════════════════════════════════════════════════════════════
# 실행
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    # 기본 시트 제거
    default = wb.active
    default.title = "_unused"

    build_revenue_sheet(wb)
    build_heatmap_sheet(wb)

    # 기본 시트 삭제
    del wb["_unused"]

    wb.save(OUT_FILE)
    print(f"✅ 저장 완료: {OUT_FILE}")
    print(f"   시트 1: 1.3_매출추이분석  (CAGR + FY Mix% 자동 계산)")
    print(f"   시트 2: 히트맵_Revenue     (파트너 × Competency 히트맵)")

if __name__ == "__main__":
    main()
